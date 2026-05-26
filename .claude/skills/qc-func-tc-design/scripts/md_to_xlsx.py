# -*- coding: utf-8 -*-
"""Convert QC test case markdown drafts (*.md, possibly multi-part) into the
test case xlsx template (`templates/Testcase_template.xlsx`).

Usage:
    python md_to_xlsx.py --input-glob "docs/QC/test-cases/functional-test/UC161-166/UC161-166_*_v2_part*.md" --uc-id UC161-166

Layout produced (single sheet "Test cases"):
    Row 1            : original column headers from template (untouched)
    Row 2+           : interleaved
                       - screen header rows  (## I. Màn hình: ...)        -> col B
                       - section header rows (### I.1./I.2. ...)          -> col B
                       - test case rows                                   -> col A..F

Column mapping (matches template "Test cases" sheet):
    A = TC ID | B = Test Title/Summary of test cases | C = Pre-conditions
    D = Test Steps | E = Expected Result | F = Priority
"""
from __future__ import annotations

import argparse
import glob
import os
import re
import shutil
import sys
from dataclasses import dataclass
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_TEMPLATE = os.path.normpath(
    os.path.join(SCRIPT_DIR, "..", "templates", "Testcase_template.xlsx")
)
SHEET_NAME = "Test cases"
EXPECTED_HEADER_FIRST_CELL = "TC ID"

PART_NUM_RE = re.compile(r"part(\d+)", re.IGNORECASE)
TABLE_HEADER_RE = re.compile(r"^\|\s*TC\s*ID\s*\|", re.IGNORECASE)
TABLE_SEP_RE = re.compile(r"^\|\s*[:\- ]+\s*\|")
TC_ROW_RE = re.compile(r"^\|\s*TC[_\-]?\d", re.IGNORECASE)
ANNOTATION_RE = re.compile(r"\s*\[(NEW|UPDATED|DELETED|MOVED)[^\]]*\]", re.IGNORECASE)
VERSION_SUFFIX_RE = re.compile(r"_v(\d+)\.xlsx$", re.IGNORECASE)
DRAFT_PART_SUFFIX_RE = re.compile(
    r"_(?:draft|draft_v\d+)?_?part\d+\.md$", re.IGNORECASE
)
VIET_DIACRITIC_RE = re.compile(
    "[àáâãèéêìíòóôõùúýăđĩũơưạảấầẩẫậắằẳẵặẹẻẽếềểễệỉịọỏốồổỗộớờởỡợụủứừửữựỳỵỷỹ"
    "ÀÁÂÃÈÉÊÌÍÒÓÔÕÙÚÝĂĐĨŨƠƯẠẢẤẦẨẪẬẮẰẲẴẶẸẺẼẾỀỂỄỆỈỊỌỎỐỒỔỖỘỚỜỞỠỢỤỦỨỪỬỮỰỲỴỶỸ]"
)
MOJIBAKE_MARKERS = ("Ã\x83", "Ã\x82", "Ä\x90", "â\x80", "\xc3\x83")


@dataclass
class TestCase:
    tc_id: str
    title: str
    pre_conditions: str
    test_steps: str
    expected_result: str
    priority: str


@dataclass
class HeaderRow:
    text: str
    level: str  # "screen" or "section"


Item = TestCase | HeaderRow


def part_number(path: str) -> int:
    m = PART_NUM_RE.search(os.path.basename(path))
    return int(m.group(1)) if m else 0


def collect_inputs(input_glob: str) -> list[str]:
    paths = sorted(glob.glob(input_glob), key=part_number)
    if not paths:
        raise SystemExit(f"No markdown files matched glob: {input_glob}")
    return paths


def split_table_row(line: str) -> list[str]:
    return [c.strip() for c in line.strip().strip("|").split("|")]


def unescape_md(text: str) -> str:
    return text.replace("\\n", "\n").replace("<br>", "\n").replace("<br/>", "\n")


def clean_title(title: str) -> str:
    return ANNOTATION_RE.sub("", title).strip()


def parse_md_files(paths: Iterable[str]) -> list[Item]:
    items: list[Item] = []
    in_table = False
    table_columns: list[str] = []

    for path in paths:
        with open(path, "r", encoding="utf-8") as fh:
            for raw in fh:
                line = raw.rstrip("\n")
                stripped = line.strip()

                if stripped.startswith("### "):
                    items.append(HeaderRow(stripped[4:].strip(), level="section"))
                    in_table = False
                    continue
                if stripped.startswith("## ") and not stripped.startswith("### "):
                    items.append(HeaderRow(stripped[3:].strip(), level="screen"))
                    in_table = False
                    continue
                if stripped.startswith("#"):
                    in_table = False
                    continue

                if not stripped.startswith("|"):
                    in_table = False
                    continue

                if TABLE_HEADER_RE.match(stripped):
                    table_columns = split_table_row(stripped)
                    in_table = True
                    continue
                if TABLE_SEP_RE.match(stripped):
                    continue
                if in_table and TC_ROW_RE.match(stripped):
                    cells = split_table_row(stripped)
                    row_dict = dict(zip(table_columns, cells))
                    items.append(
                        TestCase(
                            tc_id=row_dict.get("TC ID", "").strip(),
                            title=clean_title(row_dict.get("Title", "")),
                            pre_conditions=unescape_md(row_dict.get("Pre-conditions", "")),
                            test_steps=unescape_md(row_dict.get("Test Steps", "")),
                            expected_result=unescape_md(row_dict.get("Expected Result", "")),
                            priority=row_dict.get("Priority", "").strip(),
                        )
                    )
    return items


def derive_output_basename(first_md_path: str) -> str:
    name = os.path.basename(first_md_path)
    base = DRAFT_PART_SUFFIX_RE.sub("", name)
    if base == name:
        base = os.path.splitext(name)[0]
    return base


def next_version(output_dir: str, uc_id: str) -> int:
    """Highest existing version + 1, scanning all xlsx in output_dir whose name
    contains the UC id and ends with `_v{N}.xlsx`. Tolerates extra tokens (e.g.
    date stamps) between the UC id and the version suffix."""
    versions = []
    for path in glob.glob(os.path.join(output_dir, "*.xlsx")):
        name = os.path.basename(path)
        if uc_id not in name:
            continue
        m = VERSION_SUFFIX_RE.search(name)
        if m:
            versions.append(int(m.group(1)))
    return (max(versions) + 1) if versions else 1


def write_workbook(items: list[Item], template_path: str, output_path: str) -> tuple[int, int]:
    shutil.copy2(template_path, output_path)
    wb = load_workbook(output_path)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(
            f"Template missing expected sheet '{SHEET_NAME}'. Found sheets: {wb.sheetnames}"
        )
    ws = wb[SHEET_NAME]
    first = ws.cell(row=1, column=1).value
    if not (isinstance(first, str) and first.strip().lower() == EXPECTED_HEADER_FIRST_CELL.lower()):
        raise SystemExit(
            f"Sheet '{SHEET_NAME}' row-1 col-A is {first!r}, expected '{EXPECTED_HEADER_FIRST_CELL}'."
        )

    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    wrap_top = Alignment(wrap_text=True, vertical="top")
    bold = Font(bold=True)

    row = 2
    tc_count = 0
    header_count = 0
    for item in items:
        if isinstance(item, HeaderRow):
            cell = ws.cell(row=row, column=2, value=item.text)
            cell.font = bold
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            header_count += 1
        else:
            ws.cell(row=row, column=1, value=item.tc_id)
            ws.cell(row=row, column=2, value=item.title)
            ws.cell(row=row, column=3, value=item.pre_conditions)
            ws.cell(row=row, column=4, value=item.test_steps)
            ws.cell(row=row, column=5, value=item.expected_result)
            ws.cell(row=row, column=6, value=item.priority)
            for col in range(1, 7):
                ws.cell(row=row, column=col).alignment = wrap_top
            tc_count += 1
        row += 1

    wb.save(output_path)
    return tc_count, header_count


def verify_diacritics(output_path: str, sample_target: int = 5) -> None:
    wb = load_workbook(output_path)
    if SHEET_NAME not in wb.sheetnames:
        return
    ws = wb[SHEET_NAME]
    samples: list[tuple[int, int, str]] = []
    for r in range(2, ws.max_row + 1):
        for c in (2, 3, 4, 5):
            val = ws.cell(row=r, column=c).value
            if isinstance(val, str) and VIET_DIACRITIC_RE.search(val):
                samples.append((r, c, val))
                if len(samples) >= sample_target:
                    break
        if len(samples) >= sample_target:
            break

    if not samples:
        print("WARN: No Vietnamese-diacritic sample found to verify.")
        return

    bad = [s for s in samples if any(m in s[2] for m in MOJIBAKE_MARKERS)]
    if bad:
        print("ERROR: Possible mojibake detected in output:")
        for r, c, v in bad:
            print(f"  row {r} col {c}: {v[:120]!r}")
        raise SystemExit(2)

    print(f"OK: Vietnamese diacritics preserved across {len(samples)} sample cells.")
    for r, c, v in samples[:3]:
        print(f"  row {r} col {c}: {v[:80]}")


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--input-glob", required=True, help="Glob for md draft files (sorted by partN).")
    p.add_argument("--uc-id", required=True, help="UC identifier, e.g. UC161-166.")
    p.add_argument("--output-dir", default=None, help="Default: directory of the first md file.")
    p.add_argument("--output-name", default=None, help="Override output filename (without dir). Default: derived from first md, suffixed _v{N}.xlsx.")
    p.add_argument("--template", default=DEFAULT_TEMPLATE, help=f"Default: {DEFAULT_TEMPLATE}")
    p.add_argument("--dry-run", action="store_true", help="Parse and report counts without writing the xlsx.")
    args = p.parse_args(argv)

    inputs = collect_inputs(args.input_glob)
    print(f"Inputs ({len(inputs)} file(s), sorted by partN):")
    for path in inputs:
        print(f"  - {path}")

    items = parse_md_files(inputs)
    tc_total = sum(1 for i in items if isinstance(i, TestCase))
    header_total = sum(1 for i in items if isinstance(i, HeaderRow))
    print(f"Parsed: {tc_total} test cases, {header_total} header rows.")

    if not items:
        raise SystemExit("Parser produced 0 items — verify md files contain TC tables.")

    output_dir = args.output_dir or os.path.dirname(os.path.abspath(inputs[0]))
    os.makedirs(output_dir, exist_ok=True)

    if args.output_name:
        output_name = args.output_name
    else:
        base = derive_output_basename(inputs[0])
        version = next_version(output_dir, args.uc_id)
        output_name = f"{base}_v{version}.xlsx"

    output_path = os.path.join(output_dir, output_name)
    print(f"Output: {output_path}")

    if args.dry_run:
        print("[dry-run] No file written.")
        return 0

    if not os.path.isfile(args.template):
        raise SystemExit(f"Template not found: {args.template}")

    if os.path.exists(output_path):
        raise SystemExit(f"Refusing to overwrite existing file: {output_path}")

    tc, hdr = write_workbook(items, args.template, output_path)
    print(f"Wrote {tc} test cases + {hdr} header rows.")

    verify_diacritics(output_path)
    return 0


if __name__ == "__main__":
    sys.exit(main())
